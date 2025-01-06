import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import calendar
from typing import Dict


class Writer:
    def __init__(self, file_name: str, num_personnel: int, rows_per_person: int):
        self.schedule: Dict[str, list] | None = None
        self.weekdays = list(calendar.day_name)
        self.file_name: str = file_name
        self.num_personnel: int = num_personnel
        self.rows_per_person: int = rows_per_person

    def write(self, schedule: Dict[str, list], night_shifts: dict, day_shifts: dict):
        self.data_frame = pd.DataFrame(schedule)
        self.day_shifts = day_shifts
        self.night_shifts = night_shifts
        self._create_table()
        self._merge_person_and_shifts_cells()
        self._center_day_cells()
        self.workbook.save(self.file_name)

    def _create_table(self) -> None:
        self.data_frame.insert(0, "Personnel", "")
        self.data_frame.columns = ["Personnel"] + self.weekdays
        self.data_frame["Day"] = ""
        self.data_frame["Night"] = ""
        self.data_frame.to_excel(self.file_name, index=False)
        self.workbook = load_workbook(self.file_name)
        self.worksheet = self.workbook.active

    def _center_day_cells(self):
        for n in range(self.num_personnel):
            start_row, end_row = self._start_end_rows(n)
            for row in range(start_row, end_row + 1):  # Rows 1 to 3
                for col in range(2, 9):  # Columns A to C
                    cell = self.worksheet.cell(row=row, column=col)
                    cell.alignment = Alignment(horizontal="center", vertical="center")

    def _start_end_rows(self, n: int) -> tuple[int, int]:
        start_row = n * self.rows_per_person + 2
        end_row = start_row + self.rows_per_person - 1
        return start_row, end_row

    def _merge_person_and_shifts_cells(self) -> None:
        for n in range(self.num_personnel):
            self._merge_personnel_cells(n)
            self._merge_dayshift_cells(n)
            self._merge_nightshift_cells(n)

    def _merge_cells(self, start_row: int, end_row: int, column: int) -> None:
        self.worksheet.merge_cells(
            start_row=start_row, start_column=column, end_row=end_row, end_column=column
        )
        cell = self.worksheet.cell(row=start_row, column=column)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    def _merge_personnel_cells(self, n: int) -> None:
        start_row, end_row = self._start_end_rows(n)
        self._merge_cells(start_row=start_row, end_row=end_row, column=1)
        self.worksheet.cell(row=start_row, column=1, value=f"P{n+1}")

    def _merge_dayshift_cells(self, n: int) -> None:
        start_row, end_row = self._start_end_rows(n)
        self._merge_cells(start_row=start_row, end_row=end_row, column=9)
        self.worksheet.cell(row=start_row, column=9, value=self.day_shifts.get(n + 1))

    def _merge_nightshift_cells(self, n: int) -> None:
        start_row, end_row = self._start_end_rows(n)
        self._merge_cells(start_row=start_row, end_row=end_row, column=10)
        self.worksheet.cell(
            row=start_row, column=10, value=self.night_shifts.get(n + 1)
        )
